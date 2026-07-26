import io
import os
import sqlite3
import tempfile
import unittest
import time
from contextlib import closing
from pathlib import Path

TEST_DIR = tempfile.TemporaryDirectory()
os.environ["PACKING_DATA_DIR"] = TEST_DIR.name
os.environ["PACKING_DB_PATH"] = str(Path(TEST_DIR.name) / "test.db")

from app import BACKUP_DIR, IS_POSTGRES, app, daily_backup  # noqa: E402
from openpyxl import load_workbook  # noqa: E402


SOURCE = Path(__file__).parents[2] / "拣货单_2026-07-20_17-07-35.xlsx"


class PackingFlowTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        app.config.update(TESTING=True)
        cls.client = app.test_client()
        with cls.client.session_transaction() as current_session:
            current_session["csrf_token"] = "test-csrf-token"
        cls.csrf = {"X-CSRF-Token": "test-csrf-token"}

    def import_source(self, internal_order="测试内部单号"):
        with SOURCE.open("rb") as source:
            response = self.client.post(
                "/api/import",
                data={"internal_order": internal_order, "file": (io.BytesIO(source.read()), SOURCE.name)},
                content_type="multipart/form-data",
                headers=self.csrf,
            )
        self.assertEqual(response.status_code, 201, response.get_json())
        return response.get_json()

    def test_complete_flow(self):
        imported = self.import_source()
        batch_id = imported["batch"]["id"]
        self.assertEqual(len(imported["skus"]), 4)
        self.assertEqual(imported["summary"]["planned"], 87)
        self.assertEqual(imported["batch"]["internal_order"], "测试内部单号")

        sku = next(row for row in imported["skus"] if row["sku_code"] == "MTWT01138")
        second_sku = next(row for row in imported["skus"] if row["sku_code"] == "MTWT01105")
        payload = {
            "package_no": "2#", "clone_count": 2,
            "length_cm": 50, "width_cm": 40, "height_cm": 30, "weight_kg": 12.5,
            "items": [{"sku_id": sku["id"], "quantity": 5}, {"sku_id": second_sku["id"], "quantity": 1}],
        }
        created = self.client.post(f"/api/batches/{batch_id}/packages", json=payload, headers=self.csrf)
        self.assertEqual(created.status_code, 201, created.get_json())
        self.assertEqual(created.get_json()["created"], ["2#", "3#"])

        conflict = self.client.post(f"/api/batches/{batch_id}/packages", json=payload, headers=self.csrf)
        self.assertEqual(conflict.status_code, 400)
        self.assertIn("大包号已存在", conflict.get_json()["error"])

        over = dict(payload, package_no=4, clone_count=1, items=[{"sku_id": sku["id"], "quantity": 3}])
        warning = self.client.post(f"/api/batches/{batch_id}/packages", json=over, headers=self.csrf)
        self.assertEqual(warning.status_code, 409, warning.get_json())
        self.assertEqual(warning.get_json()["overages"][0]["after"], 13)
        over["force"] = True
        forced = self.client.post(f"/api/batches/{batch_id}/packages", json=over, headers=self.csrf)
        self.assertEqual(forced.status_code, 201, forced.get_json())

        export = self.client.get(f"/api/batches/{batch_id}/export")
        self.assertEqual(export.status_code, 200)
        book = load_workbook(io.BytesIO(export.data), data_only=True)
        self.assertEqual(book.sheetnames, ["发货清单"])
        self.assertEqual(book["发货清单"]["D1"].value, "测试内部单号")
        self.assertEqual(book["发货清单"]["A6"].value, "2#")
        self.assertIn("\n", book["发货清单"]["B6"].value)
        self.assertEqual(book["发货清单"]["C6"].value, "5\n1")
        self.assertEqual(book["发货清单"]["D6"].value, 6)
        self.assertEqual(list(book["发货清单"].merged_cells.ranges), [])
        self.assertEqual(book["发货清单"]["F2"].value, 37.5)
        self.assertEqual(book["发货清单"]["I5"].value, "体积(m³)")
        self.assertEqual(book["发货清单"]["I6"].value, 0.06)
        self.assertEqual(book["发货清单"]["H2"].value, 0.18)

        latest = forced.get_json()["data"]
        package_two = next(p for p in latest["packages"] if p["package_no"] == 2)
        changed = dict(payload, package_no=2, clone_count=99, items=[{"sku_id": sku["id"], "quantity": 2}])
        updated = self.client.put(f'/api/packages/{package_two["id"]}', json=changed, headers=self.csrf)
        self.assertEqual(updated.status_code, 200, updated.get_json())
        self.assertEqual(updated.get_json()["data"]["summary"]["packed"], 11)

        package_three = next(p for p in updated.get_json()["data"]["packages"] if p["package_no"] == 3)
        deleted = self.client.delete(f'/api/packages/{package_three["id"]}', headers=self.csrf)
        self.assertEqual(deleted.status_code, 200)
        self.assertEqual(deleted.get_json()["data"]["summary"]["packages"], 2)
        self.assertEqual(deleted.get_json()["data"]["summary"]["packed"], 5)

        if not IS_POSTGRES:
            backups = list(BACKUP_DIR.glob("packing-*.db"))
            self.assertTrue(backups)
            with closing(sqlite3.connect(backups[0])) as backup:
                self.assertGreaterEqual(backup.execute("SELECT COUNT(*) FROM batches").fetchone()[0], 1)

    def test_validation_and_atomic_clone_conflict(self):
        imported = self.import_source()
        batch_id = imported["batch"]["id"]
        sku_id = imported["skus"][0]["id"]
        base = {
            "package_no": 1, "clone_count": 1,
            "length_cm": 50, "width_cm": 40, "height_cm": 30, "weight_kg": 1.25,
            "items": [{"sku_id": sku_id, "quantity": 1}],
        }
        for field, value, expected in [
            ("package_no", "A1", "大包号"), ("length_cm", 1.5, "长"),
            ("weight_kg", 1.234, "重量"), ("items", [], "没有商品"),
            ("clone_count", 501, "最多生成500个大包"),
        ]:
            payload = dict(base)
            payload[field] = value
            response = self.client.post(f"/api/batches/{batch_id}/packages", json=payload, headers=self.csrf)
            self.assertEqual(response.status_code, 400, response.get_json())
            self.assertIn(expected, response.get_json()["error"])

        created = self.client.post(f"/api/batches/{batch_id}/packages", json=base, headers=self.csrf)
        self.assertEqual(created.status_code, 201)
        clone = dict(base, package_no=1, clone_count=3)
        conflict = self.client.post(f"/api/batches/{batch_id}/packages", json=clone, headers=self.csrf)
        self.assertEqual(conflict.status_code, 400)
        current = self.client.get(f"/api/batches/{batch_id}").get_json()
        self.assertEqual(current["summary"]["packages"], 1)

    def test_internal_order_is_required(self):
        with SOURCE.open("rb") as source:
            response = self.client.post(
                "/api/import",
                data={"file": (io.BytesIO(source.read()), SOURCE.name)},
                content_type="multipart/form-data",
                headers=self.csrf,
            )
        self.assertEqual(response.status_code, 400)
        self.assertIn("内部单号", response.get_json()["error"])

        batches = self.client.get("/api/batches").get_json()
        self.assertLessEqual(len(batches), 3)
        found = self.client.get(f'/api/batches?q={batches[0]["batch_no"]}').get_json()
        self.assertEqual(len(found), 1)

    def test_optional_dimensions_and_csrf(self):
        imported = self.import_source("可空尺寸测试")
        batch_id = imported["batch"]["id"]
        sku_id = imported["skus"][0]["id"]
        payload = {
            "package_no": 20,
            "clone_count": 1,
            "length_cm": "",
            "width_cm": None,
            "height_cm": "",
            "weight_kg": "",
            "items": [{"sku_id": sku_id, "quantity": 1}],
        }
        rejected = self.client.post(f"/api/batches/{batch_id}/packages", json=payload)
        self.assertEqual(rejected.status_code, 403)
        created = self.client.post(f"/api/batches/{batch_id}/packages", json=payload, headers=self.csrf)
        self.assertEqual(created.status_code, 201, created.get_json())
        package = created.get_json()["data"]["packages"][0]
        self.assertIsNone(package["length_cm"])
        self.assertIsNone(package["weight_kg"])
        self.assertIsNone(package["volume_m3"])

    def test_batch_ratios_mixed_package_snapshot_and_excel(self):
        imported = self.import_source("配比测试")
        batch_id = imported["batch"]["id"]
        first, second = imported["skus"][:2]
        ratio_response = self.client.post(
            f"/api/batches/{batch_id}/ratios",
            json={"items": [
                {"sku_id": first["id"], "quantity": 1},
                {"sku_id": first["id"], "quantity": 1},
                {"sku_id": second["id"], "quantity": 1},
            ]},
            headers=self.csrf,
        )
        self.assertEqual(ratio_response.status_code, 201, ratio_response.get_json())
        ratio = ratio_response.get_json()
        self.assertEqual(ratio["name"], "配比1")
        self.assertEqual(ratio["units_per_pack"], 3)
        self.assertEqual(ratio["items"][0]["quantity"], 2)

        created = self.client.post(
            f"/api/batches/{batch_id}/packages",
            json={
                "package_no": 30, "clone_count": 2,
                "entries": [
                    {"entry_type": "sku", "sku_id": first["id"], "pack_count": 1},
                    {"entry_type": "ratio", "ratio_id": ratio["id"], "pack_count": 2},
                ],
            },
            headers=self.csrf,
        )
        self.assertEqual(created.status_code, 201, created.get_json())
        self.assertEqual(created.get_json()["created"], ["30#", "31#"])
        data = created.get_json()["data"]
        self.assertEqual(data["summary"]["packed"], 14)

        package = next(row for row in data["packages"] if row["package_no"] == 30)
        detail = self.client.get(f'/api/packages/{package["id"]}').get_json()
        self.assertEqual(len(detail["entries"]), 2)
        ratio_entry = next(row for row in detail["entries"] if row["entry_type"] == "ratio")
        self.assertEqual(ratio_entry["units_per_pack"], 3)
        self.assertEqual(ratio_entry["pack_count"], 2)
        self.assertEqual(ratio_entry["total_quantity"], 6)

        changed_ratio = self.client.put(
            f'/api/ratios/{ratio["id"]}',
            json={"items": [{"sku_id": first["id"], "quantity": 1}]},
            headers=self.csrf,
        )
        self.assertEqual(changed_ratio.status_code, 200, changed_ratio.get_json())
        self.assertEqual(changed_ratio.get_json()["units_per_pack"], 1)
        snapshot = self.client.get(f'/api/packages/{package["id"]}').get_json()
        old_ratio_entry = next(row for row in snapshot["entries"] if row["entry_type"] == "ratio")
        self.assertEqual(old_ratio_entry["units_per_pack"], 3)
        self.assertIn(f'{second["display_label"]}×1', old_ratio_entry["label"])

        export = self.client.get(f"/api/batches/{batch_id}/export")
        self.assertEqual(export.status_code, 200)
        book = load_workbook(io.BytesIO(export.data), data_only=True)
        sheet = book["发货清单"]
        self.assertEqual(sheet["B5"].value, "款色尺码/配比明细")
        self.assertEqual(sheet["D5"].value, "中包件数")
        self.assertEqual(sheet["E5"].value, "中包数量")
        self.assertEqual(book.sheetnames, ["发货清单"])
        detail_lines = sheet["B6"].value.split("\n")
        self.assertEqual(detail_lines[0], first["display_label"])
        self.assertEqual(detail_lines[1], "配比1：")
        self.assertEqual(detail_lines[2], f'{first["display_label"]} ×2')
        self.assertEqual(detail_lines[3], f'{second["display_label"]} ×1')
        self.assertEqual(sheet["C6"].value.split("\n"), ["1", "", "", ""])
        self.assertEqual(sheet["D6"].value.split("\n"), ["", "3", "", ""])
        self.assertEqual(sheet["E6"].value.split("\n"), ["", "2", "", ""])
        self.assertGreaterEqual(sheet.row_dimensions[6].height, 72)
        self.assertEqual(sheet["F6"].value, 7)
        self.assertEqual(list(sheet.merged_cells.ranges), [])

        deleted_ratio = self.client.delete(f'/api/ratios/{ratio["id"]}', headers=self.csrf)
        self.assertEqual(deleted_ratio.status_code, 200)
        current_batch = self.client.get(f"/api/batches/{batch_id}").get_json()
        self.assertEqual(current_batch["ratios"], [])
        preserved = self.client.get(f'/api/packages/{package["id"]}').get_json()
        self.assertEqual(next(row for row in preserved["entries"] if row["entry_type"] == "ratio")["units_per_pack"], 3)

        deleted_package = self.client.delete(f'/api/packages/{package["id"]}', headers=self.csrf)
        self.assertEqual(deleted_package.status_code, 200)
        self.assertEqual(deleted_package.get_json()["data"]["summary"]["packed"], 7)

    def test_ratio_batch_isolation_and_validation(self):
        first_batch = self.import_source("配比隔离1")
        second_batch = self.import_source("配比隔离2")
        response = self.client.post(
            f'/api/batches/{second_batch["batch"]["id"]}/ratios',
            json={"items": [{"sku_id": first_batch["skus"][0]["id"], "quantity": 1}]},
            headers=self.csrf,
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("不属于当前批次", response.get_json()["error"])
        empty = self.client.post(
            f'/api/batches/{second_batch["batch"]["id"]}/ratios', json={"items": []}, headers=self.csrf,
        )
        self.assertEqual(empty.status_code, 400)
        self.assertIn("至少", empty.get_json()["error"])
        malformed_ratio = self.client.post(
            f'/api/batches/{second_batch["batch"]["id"]}/ratios', json={"items": ["<script>"]}, headers=self.csrf,
        )
        self.assertEqual(malformed_ratio.status_code, 400)
        malformed_package = self.client.post(
            f'/api/batches/{second_batch["batch"]["id"]}/packages', json=["not-an-object"], headers=self.csrf,
        )
        self.assertEqual(malformed_package.status_code, 400)

        local_sku = next(row for row in second_batch["skus"] if row["planned_qty"] == 12)
        ratio = self.client.post(
            f'/api/batches/{second_batch["batch"]["id"]}/ratios',
            json={"items": [{"sku_id": local_sku["id"], "quantity": 7}]}, headers=self.csrf,
        ).get_json()
        ratio_package = {
            "package_no": 50, "clone_count": 1,
            "entries": [{"entry_type": "ratio", "ratio_id": ratio["id"], "pack_count": 2}],
        }
        warning = self.client.post(
            f'/api/batches/{second_batch["batch"]["id"]}/packages', json=ratio_package, headers=self.csrf,
        )
        self.assertEqual(warning.status_code, 409)
        self.assertEqual(warning.get_json()["overages"][0]["added"], 14)
        forced = self.client.post(
            f'/api/batches/{second_batch["batch"]["id"]}/packages',
            json=ratio_package | {"force": True}, headers=self.csrf,
        )
        self.assertEqual(forced.status_code, 201, forced.get_json())

    def test_home_security_and_history_search(self):
        home = self.client.get("/")
        self.assertEqual(home.status_code, 200)
        self.assertEqual(home.headers["X-Frame-Options"], "DENY")
        self.assertIn("frame-ancestors 'none'", home.headers["Content-Security-Policy"])
        page = home.get_data(as_text=True)
        self.assertIn('id="batchSearchBtn"', page)
        self.assertIn('role="combobox"', page)
        self.assertIn('aria-controls="suggestions"', page)

        created = [self.import_source(f"历史查询-{number}") for number in range(5)]
        prefix = created[0]["batch"]["batch_no"].split("-")[0]
        history = self.client.get(f"/api/batches?q={prefix}").get_json()
        latest = self.client.get("/api/batches").get_json()
        self.assertGreaterEqual(len(history), 5)
        self.assertLessEqual(len(history), 100)
        self.assertLessEqual(len(latest), 3)

    def test_sku_sort_volume_auto_allocation_and_backup_download(self):
        imported = self.import_source("自动配比测试")
        batch_id = imported["batch"]["id"]
        self.assertEqual(
            [row["sku_code"] for row in imported["skus"]],
            ["JYMK02533", "MTWT01105", "MTWT01138", "MTWT01139"],
        )

        chosen = imported["skus"][:3]
        payload = {
            "mode": "balanced", "start_package_no": "100#", "package_count": 7,
            "selected_sku_ids": [row["id"] for row in chosen],
        }
        preview_response = self.client.post(
            f"/api/batches/{batch_id}/auto-allocation/preview", json=payload, headers=self.csrf,
        )
        self.assertEqual(preview_response.status_code, 200, preview_response.get_json())
        preview = preview_response.get_json()
        self.assertEqual(preview["unallocated"], 0)
        self.assertLessEqual(preview["max_package_quantity"] - preview["min_package_quantity"], 1)
        expected = {row["id"]: row["remaining_qty"] for row in chosen}
        actual = {}
        for package in preview["packages"]:
            for item in package["items"]:
                actual[item["sku_id"]] = actual.get(item["sku_id"], 0) + item["quantity"]
        self.assertEqual(actual, expected)

        commit = self.client.post(
            f"/api/batches/{batch_id}/auto-allocation/commit",
            json=payload | {"preview_token": preview["preview_token"]}, headers=self.csrf,
        )
        self.assertEqual(commit.status_code, 201, commit.get_json())
        self.assertEqual(commit.get_json()["created"], [f"{number}#" for number in range(100, 107)])
        self.assertEqual(commit.get_json()["data"]["summary"]["packed"], preview["selected_total"])

        last_sku = imported["skus"][3]
        stale_payload = {
            "mode": "balanced", "start_package_no": 200, "package_count": 2,
            "selected_sku_ids": [last_sku["id"]],
        }
        stale_preview = self.client.post(
            f"/api/batches/{batch_id}/auto-allocation/preview", json=stale_payload, headers=self.csrf,
        ).get_json()
        changed_inventory = self.client.post(
            f"/api/batches/{batch_id}/packages",
            json={"package_no": 90, "clone_count": 1,
                  "items": [{"sku_id": last_sku["id"], "quantity": 1}]}, headers=self.csrf,
        )
        self.assertEqual(changed_inventory.status_code, 201, changed_inventory.get_json())
        stale_commit = self.client.post(
            f"/api/batches/{batch_id}/auto-allocation/commit",
            json=stale_payload | {"preview_token": stale_preview["preview_token"]}, headers=self.csrf,
        )
        self.assertEqual(stale_commit.status_code, 409)
        self.assertIn("重新预览", stale_commit.get_json()["error"])

        conflict_preview = self.client.post(
            f"/api/batches/{batch_id}/auto-allocation/preview",
            json={"mode": "balanced", "start_package_no": 100, "package_count": 1,
                  "selected_sku_ids": [imported["skus"][3]["id"]]}, headers=self.csrf,
        ).get_json()
        conflict = self.client.post(
            f"/api/batches/{batch_id}/auto-allocation/commit",
            json={"mode": "balanced", "start_package_no": 100, "package_count": 1,
                  "selected_sku_ids": [imported["skus"][3]["id"],],
                  "preview_token": conflict_preview["preview_token"]}, headers=self.csrf,
        )
        self.assertEqual(conflict.status_code, 409)
        self.assertIn("大包号已存在", conflict.get_json()["error"])

        valid_backup = BACKUP_DIR / "packing-20990101-010101.sql.gz"
        valid_backup.write_bytes(b"backup-test")
        invalid_backup = BACKUP_DIR / "secret.txt"
        invalid_backup.write_text("no", encoding="utf-8")
        backups = self.client.get("/api/backups").get_json()
        self.assertTrue(any(row["filename"] == valid_backup.name for row in backups))
        downloaded = self.client.get(f"/api/backups/{valid_backup.name}")
        self.assertEqual(downloaded.status_code, 200)
        self.assertEqual(downloaded.data, b"backup-test")
        downloaded.close()
        self.assertEqual(self.client.get("/api/backups/secret.txt").status_code, 400)

        if not IS_POSTGRES:
            old = BACKUP_DIR / "packing-2000-01-01.db"
            old.write_bytes(b"old")
            old_time = time.time() - 9 * 24 * 60 * 60
            os.utime(old, (old_time, old_time))
            daily_backup()
            self.assertFalse(old.exists())


if __name__ == "__main__":
    unittest.main()
