from __future__ import annotations

import io
import hashlib
import os
import re
import sqlite3
import secrets
from contextlib import contextmanager
from datetime import datetime, timedelta
from pathlib import Path

from flask import Flask, jsonify, render_template, request, send_file, session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = Path(os.getenv("PACKING_DATA_DIR", BASE_DIR / "data"))
UPLOAD_DIR = DATA_DIR / "uploads"
BACKUP_DIR = DATA_DIR / "backups"
DB_PATH = Path(os.getenv("PACKING_DB_PATH", DATA_DIR / "packing.db"))
DATABASE_URL = os.getenv("DATABASE_URL", "").strip()
IS_POSTGRES = DATABASE_URL.startswith(("postgresql://", "postgres://"))

REQUIRED_HEADERS = {"商品编码", "商品名", "款式编码", "颜色规格", "数量"}
PACKAGE_RE = re.compile(r"^\s*(\d+)\s*#?\s*$")
BACKUP_RE = re.compile(r"^packing-[0-9-]+\.(?:db|sql\.gz)$")

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", secrets.token_hex(32))
app.config["MAX_CONTENT_LENGTH"] = 20 * 1024 * 1024
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = 0
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Strict"
app.config["SESSION_COOKIE_SECURE"] = os.getenv("COOKIE_SECURE", "0") == "1"


class Connection:
    """Small compatibility layer: SQLite locally, PostgreSQL in Docker."""

    def __init__(self, raw):
        self.raw = raw

    def execute(self, sql: str, params=()):
        if IS_POSTGRES:
            sql = sql.replace("?", "%s")
        return self.raw.execute(sql, params)


def insert_id(conn: Connection, sql: str, params) -> int:
    if IS_POSTGRES:
        return int(conn.execute(f"{sql} RETURNING id", params).fetchone()["id"])
    return int(conn.execute(sql, params).lastrowid)


@contextmanager
def db():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    if IS_POSTGRES:
        from psycopg import connect
        from psycopg.rows import dict_row

        raw = connect(DATABASE_URL, row_factory=dict_row, connect_timeout=10)
    else:
        raw = sqlite3.connect(DB_PATH, timeout=10)
        raw.row_factory = sqlite3.Row
        raw.execute("PRAGMA journal_mode=WAL")
        raw.execute("PRAGMA foreign_keys=ON")
    conn = Connection(raw)
    try:
        yield conn
        raw.commit()
    except Exception:
        raw.rollback()
        raise
    finally:
        raw.close()


def init_db() -> None:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    id_type = "BIGSERIAL" if IS_POSTGRES else "INTEGER"
    fk_type = "BIGINT" if IS_POSTGRES else "INTEGER"
    with db() as conn:
        statements = [
            f"""
            CREATE TABLE IF NOT EXISTS batches (
              id {id_type} PRIMARY KEY,
              batch_no TEXT NOT NULL UNIQUE,
              batch_name TEXT NOT NULL DEFAULT '',
              internal_order TEXT NOT NULL DEFAULT '',
              source_filename TEXT NOT NULL,
              created_at TEXT NOT NULL,
              updated_at TEXT NOT NULL
            )
            """,
            f"""
            CREATE TABLE IF NOT EXISTS skus (
              id {id_type} PRIMARY KEY,
              batch_id {fk_type} NOT NULL REFERENCES batches(id) ON DELETE CASCADE,
              sku_code TEXT NOT NULL,
              product_name TEXT NOT NULL,
              style_code TEXT NOT NULL,
              color_spec TEXT NOT NULL,
              erp_color_spec TEXT NOT NULL DEFAULT '',
              warehouse TEXT NOT NULL DEFAULT '',
              planned_qty INTEGER NOT NULL,
              display_label TEXT NOT NULL,
              UNIQUE(batch_id, sku_code)
            )
            """,
            f"""
            CREATE TABLE IF NOT EXISTS packages (
              id {id_type} PRIMARY KEY,
              batch_id {fk_type} NOT NULL REFERENCES batches(id) ON DELETE CASCADE,
              package_no INTEGER NOT NULL,
              length_cm INTEGER,
              width_cm INTEGER,
              height_cm INTEGER,
              weight_kg REAL,
              created_at TEXT NOT NULL,
              updated_at TEXT NOT NULL,
              UNIQUE(batch_id, package_no)
            )
            """,
            f"""
            CREATE TABLE IF NOT EXISTS package_items (
              id {id_type} PRIMARY KEY,
              package_id {fk_type} NOT NULL REFERENCES packages(id) ON DELETE CASCADE,
              sku_id {fk_type} NOT NULL REFERENCES skus(id) ON DELETE RESTRICT,
              quantity INTEGER NOT NULL,
              sort_order INTEGER NOT NULL,
              UNIQUE(package_id, sku_id)
            )
            """,
            f"""
            CREATE TABLE IF NOT EXISTS packing_ratios (
              id {id_type} PRIMARY KEY,
              batch_id {fk_type} NOT NULL REFERENCES batches(id) ON DELETE CASCADE,
              ratio_no INTEGER NOT NULL,
              name TEXT NOT NULL,
              is_active INTEGER NOT NULL DEFAULT 1,
              created_at TEXT NOT NULL,
              updated_at TEXT NOT NULL,
              UNIQUE(batch_id, ratio_no)
            )
            """,
            f"""
            CREATE TABLE IF NOT EXISTS packing_ratio_items (
              id {id_type} PRIMARY KEY,
              ratio_id {fk_type} NOT NULL REFERENCES packing_ratios(id) ON DELETE CASCADE,
              sku_id {fk_type} NOT NULL REFERENCES skus(id) ON DELETE RESTRICT,
              quantity INTEGER NOT NULL,
              sort_order INTEGER NOT NULL,
              UNIQUE(ratio_id, sku_id)
            )
            """,
            f"""
            CREATE TABLE IF NOT EXISTS package_entries (
              id {id_type} PRIMARY KEY,
              package_id {fk_type} NOT NULL REFERENCES packages(id) ON DELETE CASCADE,
              entry_type TEXT NOT NULL,
              sku_id {fk_type} REFERENCES skus(id) ON DELETE RESTRICT,
              ratio_id {fk_type} REFERENCES packing_ratios(id) ON DELETE SET NULL,
              label_snapshot TEXT NOT NULL,
              units_per_pack INTEGER NOT NULL,
              pack_count INTEGER NOT NULL,
              sort_order INTEGER NOT NULL
            )
            """,
            f"""
            CREATE TABLE IF NOT EXISTS package_entry_items (
              id {id_type} PRIMARY KEY,
              entry_id {fk_type} NOT NULL REFERENCES package_entries(id) ON DELETE CASCADE,
              sku_id {fk_type} NOT NULL REFERENCES skus(id) ON DELETE RESTRICT,
              label_snapshot TEXT NOT NULL,
              quantity_per_pack INTEGER NOT NULL,
              sort_order INTEGER NOT NULL,
              UNIQUE(entry_id, sku_id)
            )
            """,
        ]
        for statement in statements:
            conn.execute(statement)
        if IS_POSTGRES:
            conn.execute("ALTER TABLE packages ALTER COLUMN length_cm DROP NOT NULL")
            conn.execute("ALTER TABLE packages ALTER COLUMN width_cm DROP NOT NULL")
            conn.execute("ALTER TABLE packages ALTER COLUMN height_cm DROP NOT NULL")
            conn.execute("ALTER TABLE packages ALTER COLUMN weight_kg DROP NOT NULL")
        else:
            columns = {row["name"] for row in conn.execute("PRAGMA table_info(batches)")}
            if "batch_name" not in columns:
                conn.execute("ALTER TABLE batches ADD COLUMN batch_name TEXT NOT NULL DEFAULT ''")
    if not IS_POSTGRES:
        migrate_sqlite_nullable_packages()
    daily_backup()


def migrate_sqlite_nullable_packages() -> None:
    """Preserve existing Mac data while making dimensions/weight optional."""
    raw = sqlite3.connect(DB_PATH, timeout=10)
    raw.row_factory = sqlite3.Row
    try:
        columns = {row["name"]: row for row in raw.execute("PRAGMA table_info(packages)")}
        optional = ("length_cm", "width_cm", "height_cm", "weight_kg")
        if not columns or not any(columns[name]["notnull"] for name in optional):
            return
        raw.execute("PRAGMA foreign_keys=OFF")
        raw.executescript(
            """
            BEGIN;
            CREATE TABLE packages_new (
              id INTEGER PRIMARY KEY,
              batch_id INTEGER NOT NULL REFERENCES batches(id) ON DELETE CASCADE,
              package_no INTEGER NOT NULL,
              length_cm INTEGER,
              width_cm INTEGER,
              height_cm INTEGER,
              weight_kg REAL,
              created_at TEXT NOT NULL,
              updated_at TEXT NOT NULL,
              UNIQUE(batch_id, package_no)
            );
            INSERT INTO packages_new SELECT * FROM packages;
            DROP TABLE packages;
            ALTER TABLE packages_new RENAME TO packages;
            COMMIT;
            """
        )
    finally:
        raw.close()


def daily_backup() -> None:
    if IS_POSTGRES or not DB_PATH.exists():
        return
    target = BACKUP_DIR / f"packing-{datetime.now():%Y-%m-%d}.db"
    temp = target.with_suffix(".tmp")
    source = sqlite3.connect(DB_PATH)
    destination = sqlite3.connect(temp)
    try:
        source.backup(destination)
    finally:
        destination.close()
        source.close()
    temp.replace(target)
    cutoff = datetime.now() - timedelta(days=7)
    for backup in BACKUP_DIR.glob("packing-*.db"):
        if datetime.fromtimestamp(backup.stat().st_mtime) < cutoff:
            backup.unlink()


def now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def next_batch_no(conn: sqlite3.Connection) -> str:
    prefix = datetime.now().strftime("%Y%m%d")
    row = conn.execute(
        "SELECT batch_no FROM batches WHERE batch_no LIKE ? ORDER BY batch_no DESC LIMIT 1",
        (f"{prefix}-%",),
    ).fetchone()
    seq = int(row["batch_no"].split("-")[-1]) + 1 if row else 1
    return f"{prefix}-{seq:03d}"


def parse_package_no(value) -> int:
    match = PACKAGE_RE.match(str(value or ""))
    if not match or int(match.group(1)) < 1:
        raise ValueError("大包号请输入大于0的整数，系统会自动补#")
    return int(match.group(1))


def as_positive_int(value, label: str) -> int:
    try:
        number = int(value)
        if float(value) != number or number <= 0:
            raise ValueError
        return number
    except (TypeError, ValueError):
        raise ValueError(f"{label}必须是大于0的整数")


def as_weight(value) -> float:
    try:
        number = float(value)
        if number <= 0 or round(number, 2) != number:
            raise ValueError
        return number
    except (TypeError, ValueError):
        raise ValueError("重量必须大于0，最多保留两位小数")


def as_optional_positive_int(value, label: str) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    return as_positive_int(value, label)


def as_optional_weight(value) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    return as_weight(value)


def natural_key(value: object) -> tuple:
    return tuple(int(part) if part.isdigit() else part.casefold() for part in re.split(r"(\d+)", str(value or "")))


def split_color_size(value: object) -> tuple[str, str]:
    parts = [part.strip() for part in re.split(r"[;；/|]", str(value or "")) if part.strip()]
    return (" ".join(parts[:-1]), parts[-1]) if len(parts) > 1 else (str(value or "").strip(), "")


def size_key(value: object) -> tuple:
    raw = str(value or "").strip().upper().replace(" ", "")
    raw = {"2XL": "XXL", "3XL": "XXXL"}.get(raw, raw)
    order = {"XXS": 0, "XS": 1, "S": 2, "M": 3, "L": 4, "XL": 5, "XXL": 6, "XXXL": 7}
    if raw in order:
        return (0, order[raw], ())
    match = re.fullmatch(r"(\d+)XL", raw)
    if match:
        return (0, 4 + int(match.group(1)), ())
    if raw in {"均码", "F", "FREE", "ONESIZE"}:
        return (1, 0, ())
    return (2, 0, natural_key(raw))


def sku_sort_key(row) -> tuple:
    color, size = split_color_size(row["color_spec"])
    return (natural_key(row["style_code"]), natural_key(color), size_key(size), natural_key(row["sku_code"]))


def package_volume(row) -> float | None:
    values = [row[key] for key in ("length_cm", "width_cm", "height_cm")]
    if any(value is None for value in values):
        return None
    return round(float(values[0]) * float(values[1]) * float(values[2]) / 1_000_000, 6)


def package_dict(row) -> dict:
    return dict(row) | {"package_label": f'{row["package_no"]}#', "volume_m3": package_volume(row)}


def lock_batch(conn: Connection, batch_id: int) -> None:
    suffix = " FOR UPDATE" if IS_POSTGRES else ""
    if not conn.execute(f"SELECT id FROM batches WHERE id=?{suffix}", (batch_id,)).fetchone():
        raise LookupError("批次不存在")


def is_unique_violation(exc: Exception) -> bool:
    return isinstance(exc, sqlite3.IntegrityError) or getattr(exc, "sqlstate", None) == "23505"


def ratio_detail(name: str, items: list[dict]) -> str:
    details = "；".join(f'{item["label"]}×{item["quantity"]}' for item in items)
    return f"{name}：{details}"


def ratio_record(conn: Connection, ratio_id: int, active_only: bool = True) -> dict | None:
    active_sql = " AND r.is_active=1" if active_only else ""
    ratio = conn.execute(
        f"SELECT r.* FROM packing_ratios r WHERE r.id=?{active_sql}", (ratio_id,)
    ).fetchone()
    if not ratio:
        return None
    rows = conn.execute(
        """SELECT pri.sku_id,pri.quantity,pri.sort_order,s.display_label
           FROM packing_ratio_items pri JOIN skus s ON s.id=pri.sku_id
           WHERE pri.ratio_id=? ORDER BY pri.sort_order""",
        (ratio_id,),
    ).fetchall()
    items = [
        {"sku_id": row["sku_id"], "quantity": row["quantity"], "label": row["display_label"]}
        for row in rows
    ]
    result = dict(ratio)
    result["items"] = items
    result["units_per_pack"] = sum(item["quantity"] for item in items)
    result["detail"] = ratio_detail(result["name"], items)
    return result


def batch_ratios(conn: Connection, batch_id: int) -> list[dict]:
    rows = conn.execute(
        "SELECT id FROM packing_ratios WHERE batch_id=? AND is_active=1 ORDER BY ratio_no", (batch_id,)
    ).fetchall()
    return [ratio_record(conn, row["id"]) for row in rows]


def validate_ratio_items(conn: Connection, batch_id: int, payload: dict) -> list[dict]:
    if not isinstance(payload, dict):
        raise ValueError("请求数据格式不正确")
    raw_items = payload.get("items") or []
    if not isinstance(raw_items, list):
        raise ValueError("配比明细格式不正确")
    merged: dict[int, int] = {}
    order: list[int] = []
    for item in raw_items:
        if not isinstance(item, dict):
            raise ValueError("配比明细格式不正确")
        try:
            sku_id = int(item.get("sku_id", 0))
        except (TypeError, ValueError):
            sku_id = 0
        quantity = as_positive_int(item.get("quantity"), "配比商品数量")
        if sku_id not in merged:
            order.append(sku_id)
        merged[sku_id] = merged.get(sku_id, 0) + quantity
    if not merged:
        raise ValueError("配比至少需要一个款色尺码")
    placeholders = ",".join("?" for _ in merged)
    rows = conn.execute(
        f"SELECT id,batch_id,display_label FROM skus WHERE id IN ({placeholders})", tuple(merged)
    ).fetchall()
    valid = {row["id"]: row for row in rows if row["batch_id"] == batch_id}
    if len(valid) != len(merged):
        raise ValueError("配比中存在不属于当前批次的款色尺码")
    return [
        {"sku_id": sku_id, "quantity": merged[sku_id], "label": valid[sku_id]["display_label"]}
        for sku_id in order
    ]


def batch_payload(conn: sqlite3.Connection, batch_id: int) -> dict:
    batch = conn.execute("SELECT * FROM batches WHERE id=?", (batch_id,)).fetchone()
    if not batch:
        raise LookupError("批次不存在")
    skus = list(conn.execute(
        """
        SELECT s.*, COALESCE(SUM(pi.quantity), 0) packed_qty
        FROM skus s
        LEFT JOIN package_items pi ON pi.sku_id=s.id
        WHERE s.batch_id=? GROUP BY s.id
        """,
        (batch_id,),
    ).fetchall())
    skus.sort(key=sku_sort_key)
    packages = conn.execute(
        """
        SELECT p.*, COALESCE(SUM(pi.quantity),0) total_qty, COUNT(pi.id) item_count
        FROM packages p LEFT JOIN package_items pi ON pi.package_id=p.id
        WHERE p.batch_id=? GROUP BY p.id ORDER BY p.package_no
        """,
        (batch_id,),
    ).fetchall()
    planned = sum(row["planned_qty"] for row in skus)
    packed = sum(row["packed_qty"] for row in skus)
    return {
        "batch": dict(batch),
        "summary": {
            "planned": planned,
            "packed": packed,
            "remaining": sum(max(0, r["planned_qty"] - r["packed_qty"]) for r in skus),
            "over": sum(max(0, r["packed_qty"] - r["planned_qty"]) for r in skus),
            "packages": len(packages),
        },
        "skus": [dict(row) | {"remaining_qty": row["planned_qty"] - row["packed_qty"]} for row in skus],
        "ratios": batch_ratios(conn, batch_id),
        "packages": [package_dict(row) for row in packages],
    }


@app.get("/")
def home():
    token = session.setdefault("csrf_token", secrets.token_urlsafe(32))
    return render_template("index.html", csrf_token=token)


@app.before_request
def verify_csrf():
    if request.method in {"POST", "PUT", "PATCH", "DELETE"}:
        expected = session.get("csrf_token", "")
        supplied = request.headers.get("X-CSRF-Token", "")
        if not expected or not supplied or not secrets.compare_digest(expected, supplied):
            return jsonify(error="页面验证已过期，请刷新后重试"), 403


@app.after_request
def security_headers(response):
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Referrer-Policy"] = "no-referrer"
    response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=()"
    response.headers["Content-Security-Policy"] = (
        "default-src 'self'; img-src 'self' data:; style-src 'self' 'unsafe-inline'; "
        "script-src 'self' 'unsafe-inline'; connect-src 'self'; frame-ancestors 'none'"
    )
    if request.path.startswith("/api/"):
        response.headers["Cache-Control"] = "no-store"
    return response


@app.get("/api/batches")
def list_batches():
    query = request.args.get("q", "").strip()[:100]
    with db() as conn:
        where = "WHERE LOWER(b.batch_no) LIKE LOWER(?)" if query else ""
        params = (f"%{query}%",) if query else ()
        limit = 100 if query else 3
        rows = conn.execute(
            f"""
            SELECT b.*, COUNT(DISTINCT p.id) package_count,
                   COALESCE(SUM(pi.quantity),0) packed_qty
            FROM batches b LEFT JOIN packages p ON p.batch_id=b.id
            LEFT JOIN package_items pi ON pi.package_id=p.id
            {where}
            GROUP BY b.id ORDER BY b.id DESC LIMIT {limit}
            """,
            params,
        ).fetchall()
        return jsonify([dict(row) for row in rows])


@app.post("/api/import")
def import_batch():
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify(error="请选择Excel文件"), 400
    if not file.filename.lower().endswith(".xlsx"):
        return jsonify(error="只支持.xlsx文件"), 400
    internal_order = request.form.get("internal_order", "").strip()
    if not internal_order:
        return jsonify(error="请填写内部单号"), 400
    try:
        content = file.read()
        # Some Jushuitan/WPS exports have incomplete worksheet dimensions.
        # Normal mode reads those files correctly while read-only mode sees one column.
        book = load_workbook(io.BytesIO(content), read_only=False, data_only=True)
        sheet = book.active
        rows = sheet.iter_rows(values_only=True)
        headers = [str(v).strip() if v is not None else "" for v in next(rows)]
        missing = REQUIRED_HEADERS - set(headers)
        if missing:
            return jsonify(error=f"缺少字段：{'、'.join(sorted(missing))}"), 400
        pos = {name: headers.index(name) for name in headers if name}
        records: dict[str, dict] = {}
        for row_no, row in enumerate(rows, start=2):
            code = str(row[pos["商品编码"]] or "").strip()
            if not code:
                continue
            raw_qty = row[pos["数量"]]
            try:
                qty = int(raw_qty)
                if float(raw_qty) != qty or qty < 0:
                    raise ValueError
            except (TypeError, ValueError):
                return jsonify(error=f"第{row_no}行数量不是非负整数"), 400
            record = {
                "sku_code": code,
                "product_name": str(row[pos["商品名"]] or "").strip(),
                "style_code": str(row[pos["款式编码"]] or "").strip(),
                "color_spec": str(row[pos["颜色规格"]] or "").strip(),
                "erp_color_spec": str(row[pos.get("ERP颜色规格", -1)] or "").strip() if "ERP颜色规格" in pos else "",
                "warehouse": str(row[pos.get("仓位", -1)] or "").strip() if "仓位" in pos else "",
                "planned_qty": qty,
            }
            if code in records:
                old = records[code]
                identity = ("product_name", "style_code", "color_spec")
                if any(old[k] != record[k] for k in identity):
                    return jsonify(error=f"商品编码{code}出现不同商品信息"), 400
                old["planned_qty"] += qty
            else:
                record["display_label"] = " ".join(
                    filter(None, [code, record["product_name"], record["style_code"], record["color_spec"]])
                )
                records[code] = record
        if not records:
            return jsonify(error="文件中没有可导入的商品"), 400
        with db() as conn:
            stamp = now()
            batch_no = next_batch_no(conn)
            batch_id = insert_id(
                conn,
                "INSERT INTO batches(batch_no,batch_name,internal_order,source_filename,created_at,updated_at) VALUES(?,?,?,?,?,?)",
                (batch_no, batch_no, internal_order, file.filename, stamp, stamp),
            )
            for record in records.values():
                conn.execute(
                    """INSERT INTO skus(batch_id,sku_code,product_name,style_code,color_spec,
                       erp_color_spec,warehouse,planned_qty,display_label) VALUES(?,?,?,?,?,?,?,?,?)""",
                    (batch_id, *record.values()),
                )
            safe_name = f"{batch_no}-{Path(file.filename).name}"
            (UPLOAD_DIR / safe_name).write_bytes(content)
            result = batch_payload(conn, batch_id)
        daily_backup()
        return jsonify(result), 201
    except Exception as exc:
        app.logger.exception("import failed")
        return jsonify(error=f"导入失败：{exc}"), 400


@app.get("/api/batches/<int:batch_id>")
def get_batch(batch_id):
    with db() as conn:
        try:
            return jsonify(batch_payload(conn, batch_id))
        except LookupError as exc:
            return jsonify(error=str(exc)), 404


@app.post("/api/batches/<int:batch_id>/ratios")
def create_ratio(batch_id):
    payload = request.get_json(force=True)
    try:
        with db() as conn:
            if not conn.execute("SELECT id FROM batches WHERE id=?", (batch_id,)).fetchone():
                raise LookupError("批次不存在")
            items = validate_ratio_items(conn, batch_id, payload)
            row = conn.execute(
                "SELECT COALESCE(MAX(ratio_no),0) next_no FROM packing_ratios WHERE batch_id=?", (batch_id,)
            ).fetchone()
            ratio_no = int(row["next_no"]) + 1
            stamp = now()
            ratio_id = insert_id(
                conn,
                """INSERT INTO packing_ratios(batch_id,ratio_no,name,is_active,created_at,updated_at)
                   VALUES(?,?,?,1,?,?)""",
                (batch_id, ratio_no, f"配比{ratio_no}", stamp, stamp),
            )
            for order, item in enumerate(items):
                conn.execute(
                    "INSERT INTO packing_ratio_items(ratio_id,sku_id,quantity,sort_order) VALUES(?,?,?,?)",
                    (ratio_id, item["sku_id"], item["quantity"], order),
                )
            conn.execute("UPDATE batches SET updated_at=? WHERE id=?", (stamp, batch_id))
            result = ratio_record(conn, ratio_id)
        daily_backup()
        return jsonify(result), 201
    except LookupError as exc:
        return jsonify(error=str(exc)), 404
    except ValueError as exc:
        return jsonify(error=str(exc)), 400


@app.get("/api/ratios/<int:ratio_id>")
def get_ratio(ratio_id):
    with db() as conn:
        result = ratio_record(conn, ratio_id)
        if not result:
            return jsonify(error="配比不存在"), 404
        return jsonify(result)


@app.put("/api/ratios/<int:ratio_id>")
def update_ratio(ratio_id):
    payload = request.get_json(force=True)
    try:
        with db() as conn:
            ratio = ratio_record(conn, ratio_id)
            if not ratio:
                return jsonify(error="配比不存在"), 404
            items = validate_ratio_items(conn, ratio["batch_id"], payload)
            conn.execute("DELETE FROM packing_ratio_items WHERE ratio_id=?", (ratio_id,))
            for order, item in enumerate(items):
                conn.execute(
                    "INSERT INTO packing_ratio_items(ratio_id,sku_id,quantity,sort_order) VALUES(?,?,?,?)",
                    (ratio_id, item["sku_id"], item["quantity"], order),
                )
            stamp = now()
            conn.execute("UPDATE packing_ratios SET updated_at=? WHERE id=?", (stamp, ratio_id))
            conn.execute("UPDATE batches SET updated_at=? WHERE id=?", (stamp, ratio["batch_id"]))
            result = ratio_record(conn, ratio_id)
        daily_backup()
        return jsonify(result)
    except ValueError as exc:
        return jsonify(error=str(exc)), 400


@app.delete("/api/ratios/<int:ratio_id>")
def delete_ratio(ratio_id):
    with db() as conn:
        ratio = ratio_record(conn, ratio_id)
        if not ratio:
            return jsonify(error="配比不存在"), 404
        stamp = now()
        conn.execute("UPDATE packing_ratios SET is_active=0,updated_at=? WHERE id=?", (stamp, ratio_id))
        conn.execute("UPDATE batches SET updated_at=? WHERE id=?", (stamp, ratio["batch_id"]))
    daily_backup()
    return jsonify(ok=True)


def fallback_package_entries(conn: Connection, package_id: int) -> list[dict]:
    rows = conn.execute(
        """SELECT pi.sku_id,pi.quantity,s.display_label FROM package_items pi
           JOIN skus s ON s.id=pi.sku_id WHERE pi.package_id=? ORDER BY pi.sort_order""",
        (package_id,),
    ).fetchall()
    return [
        {
            "entry_id": None, "entry_type": "sku", "sku_id": row["sku_id"], "ratio_id": None,
            "label": row["display_label"], "units_per_pack": 1, "pack_count": row["quantity"],
            "total_quantity": row["quantity"],
            "items": [{"sku_id": row["sku_id"], "label": row["display_label"], "quantity_per_pack": 1}],
        }
        for row in rows
    ]


def package_entries_payload(conn: Connection, package_id: int) -> list[dict]:
    rows = conn.execute(
        "SELECT * FROM package_entries WHERE package_id=? ORDER BY sort_order", (package_id,)
    ).fetchall()
    if not rows:
        return fallback_package_entries(conn, package_id)
    entries = []
    flattened: dict[int, int] = {}
    for row in rows:
        components = conn.execute(
            """SELECT pei.sku_id,pei.label_snapshot,pei.quantity_per_pack FROM package_entry_items pei
               WHERE pei.entry_id=? ORDER BY pei.sort_order""",
            (row["id"],),
        ).fetchall()
        items = []
        for item in components:
            total = item["quantity_per_pack"] * row["pack_count"]
            flattened[item["sku_id"]] = flattened.get(item["sku_id"], 0) + total
            items.append({
                "sku_id": item["sku_id"], "label": item["label_snapshot"],
                "quantity_per_pack": item["quantity_per_pack"],
            })
        entries.append({
            "entry_id": row["id"], "entry_type": row["entry_type"], "sku_id": row["sku_id"],
            "ratio_id": row["ratio_id"], "label": row["label_snapshot"],
            "units_per_pack": row["units_per_pack"], "pack_count": row["pack_count"],
            "total_quantity": row["units_per_pack"] * row["pack_count"], "items": items,
        })
    authoritative = {
        row["sku_id"]: row["quantity"] for row in conn.execute(
            "SELECT sku_id,quantity FROM package_items WHERE package_id=?", (package_id,)
        ).fetchall()
    }
    return entries if flattened == authoritative else fallback_package_entries(conn, package_id)


def resolve_package_entries(
    conn: Connection, batch_id: int, payload: dict, editing_id: int | None = None
) -> tuple[list[dict], dict[int, int]]:
    raw_entries = payload.get("entries")
    if raw_entries is None:
        legacy_items = payload.get("items") or []
        if not isinstance(legacy_items, list) or any(not isinstance(item, dict) for item in legacy_items):
            raise ValueError("大包商品明细格式不正确")
        raw_entries = [
            {"entry_type": "sku", "sku_id": item.get("sku_id"), "pack_count": item.get("quantity")}
            for item in legacy_items
        ]
    if not isinstance(raw_entries, list) or any(not isinstance(entry, dict) for entry in raw_entries):
        raise ValueError("大包商品明细格式不正确")
    resolved: list[dict] = []
    for raw_entry in raw_entries:
        pack_count = as_positive_int(raw_entry.get("pack_count"), "数量")
        entry_id = raw_entry.get("entry_id")
        if entry_id and editing_id:
            stored = conn.execute(
                "SELECT * FROM package_entries WHERE id=? AND package_id=?", (entry_id, editing_id)
            ).fetchone()
            if not stored:
                raise ValueError("大包内存在无效配比快照")
            components = conn.execute(
                "SELECT * FROM package_entry_items WHERE entry_id=? ORDER BY sort_order", (entry_id,)
            ).fetchall()
            resolved.append({
                "entry_type": stored["entry_type"], "sku_id": stored["sku_id"],
                "ratio_id": stored["ratio_id"], "label": stored["label_snapshot"],
                "units_per_pack": stored["units_per_pack"], "pack_count": pack_count,
                "items": [{
                    "sku_id": item["sku_id"], "label": item["label_snapshot"],
                    "quantity_per_pack": item["quantity_per_pack"],
                } for item in components],
            })
            continue
        entry_type = str(raw_entry.get("entry_type") or "sku")
        if entry_type == "ratio":
            try:
                ratio_id = int(raw_entry.get("ratio_id", 0))
            except (TypeError, ValueError):
                ratio_id = 0
            ratio = ratio_record(conn, ratio_id)
            if not ratio or ratio["batch_id"] != batch_id:
                raise ValueError("选择的配比不属于当前批次或已删除")
            resolved.append({
                "entry_type": "ratio", "sku_id": None, "ratio_id": ratio_id,
                "label": ratio["detail"], "units_per_pack": ratio["units_per_pack"],
                "pack_count": pack_count,
                "items": [{
                    "sku_id": item["sku_id"], "label": item["label"],
                    "quantity_per_pack": item["quantity"],
                } for item in ratio["items"]],
            })
        elif entry_type == "sku":
            try:
                sku_id = int(raw_entry.get("sku_id", 0))
            except (TypeError, ValueError):
                sku_id = 0
            sku = conn.execute(
                "SELECT id,batch_id,display_label FROM skus WHERE id=?", (sku_id,)
            ).fetchone()
            if not sku or sku["batch_id"] != batch_id:
                raise ValueError("大包内存在不属于当前批次的商品")
            resolved.append({
                "entry_type": "sku", "sku_id": sku_id, "ratio_id": None,
                "label": sku["display_label"], "units_per_pack": 1, "pack_count": pack_count,
                "items": [{"sku_id": sku_id, "label": sku["display_label"], "quantity_per_pack": 1}],
            })
        else:
            raise ValueError("大包内存在无效的商品类型")
    if not resolved:
        raise ValueError("当前大包还没有商品")
    flattened: dict[int, int] = {}
    for entry in resolved:
        for item in entry["items"]:
            quantity = item["quantity_per_pack"] * entry["pack_count"]
            flattened[item["sku_id"]] = flattened.get(item["sku_id"], 0) + quantity
    return resolved, flattened


def write_package_contents(conn: Connection, package_id: int, entries: list[dict], flattened: dict[int, int]) -> None:
    conn.execute("DELETE FROM package_entries WHERE package_id=?", (package_id,))
    conn.execute("DELETE FROM package_items WHERE package_id=?", (package_id,))
    for order, entry in enumerate(entries):
        entry_id = insert_id(
            conn,
            """INSERT INTO package_entries(package_id,entry_type,sku_id,ratio_id,label_snapshot,
               units_per_pack,pack_count,sort_order) VALUES(?,?,?,?,?,?,?,?)""",
            (package_id, entry["entry_type"], entry["sku_id"], entry["ratio_id"], entry["label"],
             entry["units_per_pack"], entry["pack_count"], order),
        )
        for item_order, item in enumerate(entry["items"]):
            conn.execute(
                """INSERT INTO package_entry_items(entry_id,sku_id,label_snapshot,quantity_per_pack,sort_order)
                   VALUES(?,?,?,?,?)""",
                (entry_id, item["sku_id"], item["label"], item["quantity_per_pack"], item_order),
            )
    for order, (sku_id, quantity) in enumerate(flattened.items()):
        conn.execute(
            "INSERT INTO package_items(package_id,sku_id,quantity,sort_order) VALUES(?,?,?,?)",
            (package_id, sku_id, quantity, order),
        )


def validate_package(conn, batch_id: int, payload: dict, editing_id: int | None = None):
    if not isinstance(payload, dict):
        raise ValueError("请求数据格式不正确")
    start_no = parse_package_no(payload.get("package_no"))
    clone_count = as_positive_int(payload.get("clone_count", 1), "同款大包数")
    if clone_count > 500:
        raise ValueError("同款大包数一次最多生成500个大包")
    length_cm = as_optional_positive_int(payload.get("length_cm"), "长")
    width_cm = as_optional_positive_int(payload.get("width_cm"), "宽")
    height_cm = as_optional_positive_int(payload.get("height_cm"), "高")
    weight_kg = as_optional_weight(payload.get("weight_kg"))
    entries, merged = resolve_package_entries(conn, batch_id, payload, editing_id)
    valid = conn.execute(
        f"SELECT id,display_label,planned_qty FROM skus WHERE batch_id=? AND id IN ({','.join('?' * len(merged))})",
        (batch_id, *merged.keys()),
    ).fetchall()
    if len(valid) != len(merged):
        raise ValueError("大包内存在不属于当前批次的商品")
    target_nos = list(range(start_no, start_no + clone_count))
    placeholders = ",".join("?" for _ in target_nos)
    params = [batch_id, *target_nos]
    sql = f"SELECT package_no FROM packages WHERE batch_id=? AND package_no IN ({placeholders})"
    if editing_id:
        sql += " AND id<>?"
        params.append(editing_id)
    conflicts = [f'{r["package_no"]}#' for r in conn.execute(sql, params).fetchall()]
    if conflicts:
        raise ValueError(f"大包号已存在：{'、'.join(conflicts)}")
    current_sql = """SELECT pi.sku_id,SUM(pi.quantity) qty FROM package_items pi
                     JOIN packages p ON p.id=pi.package_id WHERE p.batch_id=?"""
    current_params = [batch_id]
    if editing_id is not None:
        current_sql += " AND p.id<>?"
        current_params.append(editing_id)
    current_sql += " GROUP BY pi.sku_id"
    current = {r["sku_id"]: r["qty"] for r in conn.execute(current_sql, current_params).fetchall()}
    overages = []
    for sku in valid:
        after = current.get(sku["id"], 0) + merged[sku["id"]] * clone_count
        if after > sku["planned_qty"]:
            overages.append({
                "label": sku["display_label"], "planned": sku["planned_qty"],
                "before": current.get(sku["id"], 0), "added": merged[sku["id"]] * clone_count,
                "after": after, "over": after - sku["planned_qty"],
            })
    return {
        "start_no": start_no, "clone_count": clone_count, "target_nos": target_nos,
        "length_cm": length_cm, "width_cm": width_cm, "height_cm": height_cm,
        "weight_kg": weight_kg, "items": merged, "entries": entries, "overages": overages,
    }


@app.post("/api/batches/<int:batch_id>/packages")
def create_package(batch_id):
    payload = request.get_json(force=True)
    try:
        with db() as conn:
            lock_batch(conn, batch_id)
            checked = validate_package(conn, batch_id, payload)
            if checked["overages"] and not payload.get("force"):
                return jsonify(error="保存后将超过清单数量", overages=checked["overages"]), 409
            stamp = now()
            for package_no in checked["target_nos"]:
                package_id = insert_id(
                    conn,
                    """INSERT INTO packages(batch_id,package_no,length_cm,width_cm,height_cm,weight_kg,created_at,updated_at)
                       VALUES(?,?,?,?,?,?,?,?)""",
                    (batch_id, package_no, checked["length_cm"], checked["width_cm"], checked["height_cm"], checked["weight_kg"], stamp, stamp),
                )
                write_package_contents(conn, package_id, checked["entries"], checked["items"])
            conn.execute("UPDATE batches SET updated_at=? WHERE id=?", (stamp, batch_id))
            result = batch_payload(conn, batch_id)
        daily_backup()
        return jsonify(created=[f"{n}#" for n in checked["target_nos"]], data=result), 201
    except (ValueError, LookupError) as exc:
        return jsonify(error=str(exc)), 400
    except Exception as exc:
        if is_unique_violation(exc):
            return jsonify(error="大包号已存在，请刷新批次后重试"), 409
        raise


@app.get("/api/packages/<int:package_id>")
def get_package(package_id):
    with db() as conn:
        package = conn.execute("SELECT * FROM packages WHERE id=?", (package_id,)).fetchone()
        if not package:
            return jsonify(error="大包不存在"), 404
        items = conn.execute("SELECT sku_id,quantity FROM package_items WHERE package_id=? ORDER BY sort_order", (package_id,)).fetchall()
        return jsonify(package_dict(package) | {
            "items": [dict(r) for r in items],
            "entries": package_entries_payload(conn, package_id),
        })


@app.put("/api/packages/<int:package_id>")
def update_package(package_id):
    payload = request.get_json(force=True)
    try:
        with db() as conn:
            original = conn.execute("SELECT * FROM packages WHERE id=?", (package_id,)).fetchone()
            if not original:
                return jsonify(error="大包不存在"), 404
            lock_batch(conn, original["batch_id"])
            payload["clone_count"] = 1
            checked = validate_package(conn, original["batch_id"], payload, package_id)
            if checked["overages"] and not payload.get("force"):
                return jsonify(error="保存后将超过清单数量", overages=checked["overages"]), 409
            stamp = now()
            conn.execute(
                "UPDATE packages SET package_no=?,length_cm=?,width_cm=?,height_cm=?,weight_kg=?,updated_at=? WHERE id=?",
                (checked["start_no"], checked["length_cm"], checked["width_cm"], checked["height_cm"], checked["weight_kg"], stamp, package_id),
            )
            write_package_contents(conn, package_id, checked["entries"], checked["items"])
            conn.execute("UPDATE batches SET updated_at=? WHERE id=?", (stamp, original["batch_id"]))
            result = batch_payload(conn, original["batch_id"])
        daily_backup()
        return jsonify(data=result)
    except (ValueError, LookupError) as exc:
        return jsonify(error=str(exc)), 400


@app.delete("/api/packages/<int:package_id>")
def delete_package(package_id):
    with db() as conn:
        row = conn.execute("SELECT batch_id FROM packages WHERE id=?", (package_id,)).fetchone()
        if not row:
            return jsonify(error="大包不存在"), 404
        lock_batch(conn, row["batch_id"])
        conn.execute("DELETE FROM packages WHERE id=?", (package_id,))
        conn.execute("UPDATE batches SET updated_at=? WHERE id=?", (now(), row["batch_id"]))
        result = batch_payload(conn, row["batch_id"])
    daily_backup()
    return jsonify(data=result)


def allocation_preview(conn: Connection, batch_id: int, payload: dict) -> dict:
    if not isinstance(payload, dict):
        raise ValueError("请求数据格式不正确")
    if payload.get("mode", "balanced") != "balanced":
        raise ValueError("暂不支持该分配模式")
    start_no = parse_package_no(payload.get("start_package_no"))
    package_count = as_positive_int(payload.get("package_count"), "大包数量")
    if package_count > 500:
        raise ValueError("一次最多生成500个大包")
    raw_ids = payload.get("selected_sku_ids")
    if not isinstance(raw_ids, list) or not raw_ids:
        raise ValueError("请至少选择一个款色尺码")
    try:
        selected_ids = list(dict.fromkeys(int(value) for value in raw_ids))
    except (TypeError, ValueError):
        raise ValueError("款色尺码选择不正确")
    placeholders = ",".join("?" for _ in selected_ids)
    rows = list(conn.execute(
        f"""SELECT s.*,s.planned_qty-COALESCE(SUM(pi.quantity),0) remaining_qty
            FROM skus s LEFT JOIN package_items pi ON pi.sku_id=s.id
            WHERE s.batch_id=? AND s.id IN ({placeholders}) GROUP BY s.id""",
        (batch_id, *selected_ids),
    ).fetchall())
    if len(rows) != len(selected_ids):
        raise ValueError("选择中存在不属于当前批次的款色尺码")
    rows.sort(key=sku_sort_key)
    selected = [row for row in rows if row["remaining_qty"] > 0]
    if len(selected) != len(rows):
        raise ValueError("所选款色尺码中有库存已经用完，请刷新后重选")
    selected_total = sum(row["remaining_qty"] for row in selected)
    if package_count > selected_total:
        raise ValueError(f"大包数量不能超过所选剩余件数 {selected_total}")

    totals = [0] * package_count
    package_items: list[dict[int, int]] = [dict() for _ in range(package_count)]
    cursor = 0
    for sku in selected:
        quantity = int(sku["remaining_qty"])
        base, remainder = divmod(quantity, package_count)
        if base:
            for index in range(package_count):
                package_items[index][sku["id"]] = base
                totals[index] += base
        for _ in range(remainder):
            index = min(range(package_count), key=lambda i: (totals[i], (i - cursor) % package_count))
            package_items[index][sku["id"]] = package_items[index].get(sku["id"], 0) + 1
            totals[index] += 1
            cursor = (index + 1) % package_count

    labels = {row["id"]: row["display_label"] for row in selected}
    packages = [{
        "package_no": start_no + index,
        "package_label": f"{start_no + index}#",
        "total_quantity": totals[index],
        "items": [{"sku_id": sku_id, "label": labels[sku_id], "quantity": quantity}
                  for sku_id, quantity in package_items[index].items()],
    } for index in range(package_count)]
    signature_source = "|".join([
        str(batch_id), "balanced", str(start_no), str(package_count),
        *(f'{row["id"]}:{row["remaining_qty"]}' for row in selected),
    ])
    return {
        "mode": "balanced", "start_package_no": start_no, "package_count": package_count,
        "selected_total": selected_total, "unallocated": selected_total - sum(totals),
        "min_package_quantity": min(totals), "max_package_quantity": max(totals),
        "preview_token": hashlib.sha256(signature_source.encode()).hexdigest(), "packages": packages,
    }


@app.post("/api/batches/<int:batch_id>/auto-allocation/preview")
def preview_auto_allocation(batch_id):
    try:
        with db() as conn:
            if not conn.execute("SELECT id FROM batches WHERE id=?", (batch_id,)).fetchone():
                return jsonify(error="批次不存在"), 404
            return jsonify(allocation_preview(conn, batch_id, request.get_json(force=True)))
    except (ValueError, LookupError) as exc:
        return jsonify(error=str(exc)), 400


@app.post("/api/batches/<int:batch_id>/auto-allocation/commit")
def commit_auto_allocation(batch_id):
    payload = request.get_json(force=True)
    try:
        with db() as conn:
            lock_batch(conn, batch_id)
            preview = allocation_preview(conn, batch_id, payload)
            if not secrets.compare_digest(str(payload.get("preview_token") or ""), preview["preview_token"]):
                return jsonify(error="库存或参数已变化，请重新预览后再生成"), 409
            target_nos = [row["package_no"] for row in preview["packages"]]
            placeholders = ",".join("?" for _ in target_nos)
            conflicts = conn.execute(
                f"SELECT package_no FROM packages WHERE batch_id=? AND package_no IN ({placeholders})",
                (batch_id, *target_nos),
            ).fetchall()
            if conflicts:
                labels = "、".join(f'{row["package_no"]}#' for row in conflicts)
                return jsonify(error=f"大包号已存在：{labels}"), 409
            stamp = now()
            for package in preview["packages"]:
                package_id = insert_id(
                    conn,
                    """INSERT INTO packages(batch_id,package_no,length_cm,width_cm,height_cm,weight_kg,created_at,updated_at)
                       VALUES(?,?,?,?,?,?,?,?)""",
                    (batch_id, package["package_no"], None, None, None, None, stamp, stamp),
                )
                entries = [{
                    "entry_type": "sku", "sku_id": item["sku_id"], "ratio_id": None,
                    "label": item["label"], "units_per_pack": 1, "pack_count": item["quantity"],
                    "items": [{"sku_id": item["sku_id"], "label": item["label"], "quantity_per_pack": 1}],
                } for item in package["items"]]
                flattened = {item["sku_id"]: item["quantity"] for item in package["items"]}
                write_package_contents(conn, package_id, entries, flattened)
            conn.execute("UPDATE batches SET updated_at=? WHERE id=?", (stamp, batch_id))
            result = batch_payload(conn, batch_id)
        daily_backup()
        return jsonify(created=[row["package_label"] for row in preview["packages"]], data=result), 201
    except (ValueError, LookupError) as exc:
        return jsonify(error=str(exc)), 400
    except Exception as exc:
        if is_unique_violation(exc):
            return jsonify(error="大包号已存在，请刷新批次并重新预览"), 409
        raise


@app.get("/api/backups")
def list_backups():
    rows = []
    for path in BACKUP_DIR.iterdir():
        if path.is_file() and BACKUP_RE.fullmatch(path.name):
            rows.append({"filename": path.name, "size": path.stat().st_size,
                         "modified_at": datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds")})
    return jsonify(sorted(rows, key=lambda row: row["modified_at"], reverse=True)[:7])


@app.get("/api/backups/<path:filename>")
def download_backup(filename):
    if not BACKUP_RE.fullmatch(filename) or Path(filename).name != filename:
        return jsonify(error="备份文件名无效"), 400
    path = (BACKUP_DIR / filename).resolve()
    if path.parent != BACKUP_DIR.resolve() or not path.is_file():
        return jsonify(error="备份不存在"), 404
    return send_file(path, as_attachment=True, download_name=filename, mimetype="application/octet-stream")


@app.get("/api/batches/<int:batch_id>/export")
def export_batch(batch_id):
    with db() as conn:
        data = batch_payload(conn, batch_id)
        batch = data["batch"]
        packages = conn.execute("SELECT * FROM packages WHERE batch_id=? ORDER BY package_no", (batch_id,)).fetchall()
        book = Workbook()
        ws = book.active
        ws.title = "发货清单"
        total_weight = sum(float(p["weight_kg"]) for p in packages if p["weight_kg"] is not None)
        total_volume = sum(package_volume(p) or 0 for p in packages)
        ws.append(["批次号", batch["batch_no"], "内部单号", batch["internal_order"], "导出时间", datetime.now().strftime("%Y-%m-%d %H:%M")])
        ws.append(["总大包数", data["summary"]["packages"], "总件数", data["summary"]["packed"], "总重量(kg)", round(total_weight, 2), "总体积(m³)", round(total_volume, 6)])
        ws.append(["原文件", batch["source_filename"]])
        ws.append([])
        has_ratios = bool(conn.execute(
            """SELECT pe.id FROM package_entries pe JOIN packages p ON p.id=pe.package_id
               WHERE p.batch_id=? AND pe.entry_type='ratio' LIMIT 1""", (batch_id,)
        ).fetchone())
        headers = (
            ["大包号", "款色尺码/配比明细", "数量", "中包件数", "中包数量", "总数量", "长(cm)", "宽(cm)", "高(cm)", "重量(kg)", "体积(m³)"]
            if has_ratios else
            ["大包号", "款色尺码", "数量", "总数量", "长(cm)", "宽(cm)", "高(cm)", "重量(kg)", "体积(m³)"]
        )
        ws.append(headers)
        for package in packages:
            if has_ratios:
                items = package_entries_payload(conn, package["id"])
                detail_lines: list[str] = []
                quantity_lines: list[str] = []
                middle_unit_lines: list[str] = []
                middle_count_lines: list[str] = []
                for item in items:
                    if item["entry_type"] == "ratio":
                        ratio_name = str(item["label"]).partition("：")[0]
                        component_lines = [
                            f'{component["label"]} ×{component["quantity_per_pack"]}'
                            for component in item["items"]
                        ]
                        line_count = 1 + len(component_lines)
                        detail_lines.extend([f"{ratio_name}：", *component_lines])
                        quantity_lines.extend([""] * line_count)
                        middle_unit_lines.extend([str(item["units_per_pack"]), *([""] * (line_count - 1))])
                        middle_count_lines.extend([str(item["pack_count"]), *([""] * (line_count - 1))])
                    else:
                        detail_lines.append(item["label"])
                        quantity_lines.append(str(item["pack_count"]))
                        middle_unit_lines.append("")
                        middle_count_lines.append("")
                total = sum(item["total_quantity"] for item in items)
                column_text = lambda lines: "\n".join(lines) if any(lines) else None
                ws.append([
                    f'{package["package_no"]}#', "\n".join(detail_lines), column_text(quantity_lines),
                    column_text(middle_unit_lines), column_text(middle_count_lines), total,
                    package["length_cm"], package["width_cm"], package["height_cm"], package["weight_kg"], package_volume(package),
                ])
                ws.row_dimensions[ws.max_row].height = 18 * max(1, len(detail_lines))
            else:
                items = conn.execute(
                    """SELECT s.display_label,pi.quantity FROM package_items pi JOIN skus s ON s.id=pi.sku_id
                       WHERE pi.package_id=? ORDER BY pi.sort_order""", (package["id"],)
                ).fetchall()
                total = sum(item["quantity"] for item in items)
                ws.append([
                    f'{package["package_no"]}#', "\n".join(item["display_label"] for item in items),
                    "\n".join(str(item["quantity"]) for item in items), total,
                    package["length_cm"], package["width_cm"], package["height_cm"], package["weight_kg"], package_volume(package),
                ])
                ws.row_dimensions[ws.max_row].height = 18 * max(1, len(items))
        navy, thin = "17324D", Side(style="thin", color="D6DEE6")
        for cell in ws[5]:
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in ws.iter_rows(min_row=6):
            for cell in row:
                cell.border = Border(bottom=thin)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.freeze_panes = "A6"
        ws.auto_filter.ref = ws.dimensions
        widths = [12, 68, 10, 12, 12, 12, 10, 10, 10, 12, 14] if has_ratios else [12, 55, 10, 10, 10, 10, 10, 12, 14]
        for idx, width in enumerate(widths, 1): ws.column_dimensions[chr(64 + idx)].width = width
        for column, width in {"C": 12, "D": 18, "E": 12, "F": 18, "G": 14, "H": 14}.items():
            ws.column_dimensions[column].width = max(ws.column_dimensions[column].width or 0, width)
        volume_column = 11 if has_ratios else 9
        for row in ws.iter_rows(min_row=6, min_col=volume_column, max_col=volume_column):
            row[0].number_format = "0.000000"
        output = io.BytesIO()
        book.save(output)
        output.seek(0)
        filename = f"发货清单_{batch['batch_no']}_{datetime.now():%Y%m%d-%H%M%S}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


init_db()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", "8080")), debug=True)
